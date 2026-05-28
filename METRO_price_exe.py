import requests
import json
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage



url = "https://searchgw.dmall.com/app/new/search/wareSearch/v1"


headers = {
    "Host": "searchgw.dmall.com",
    "grayvenderid": "70121",
    "ticketname": "20DD646844A369AAB7D3D8A4D9539B51007DAF3F0FB9B8751E944C823DD0202EA2A1DEB233FB548CEC3F63BAD0FE4DA1C11DA0778CF66FE25F88E8F0DF4717D7B2B105B5D267B700B1D86F04777867CD8418A1FC6D523B40AC2060713BA9DF03FC87A530B0E090C7531E7E79293F9A8EEDA3C0B8332F601188CBE8D618E566AB",
    "dmtenantid": "2",
    "channel": "miniprograms",
    "appversion": "6.0.2",
    "applettype": "35",
    "platformstoregroupkey": "1dd098cb711e6e061f0fe326c3c86276@NTg5NjEtNTk0Njgx",
    "latitude": "28.253075629340277",
    "areaid": "430105",
    "v": "v6.0.2",
    "storeid": "721521",
    "graystoreid": "721521",
    "platform": "9",
    "token": "97c0cf17-a53a-43be-8c25-41c4cd408731",
    "businesscode": "1",
    "networktype": "0",
    "originbusinessformat": "1",
    "uniquecode": "mdl_wx",
    "storegroupkey": "be82e109d3fefb8651290e64831424cd@MS03MjE1MjEtNzAxMjEtSlNE",
    "deliverylng": "113.019284",
    "venderid": "70121",
    "metroidentity": "C",
    "longitude": "113.01975070529514",
    "content-type": "application/x-www-form-urlencoded",
    "deliverylat": "28.252600",
    "userid": "643560759",
    "user-agent": (
        "Mozilla/5.0 (iPhone; CPU iPhone OS 16_7_15 like Mac OS X) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) "
        "Mobile/15E148 MicroMessenger/8.0.71(0x18004730) "
        "NetType/WIFI Language/zh_CN"
    ),
    "referer": "https://servicewechat.com/wx2c9675578d432bee/56/page-frame.html",
}



def extract_goods(obj):
    """
    自动递归查找商品列表
    """
    if isinstance(obj, dict):

        for key in [
            "wareList",
            "list",
            "records",
            "items",
            "searchResult"
        ]:
            if key in obj and isinstance(obj[key], list):
                return obj[key]

        for v in obj.values():
            result = extract_goods(v)
            if result:
                return result

    return []

# =========================
# 分 -> 元
# =========================
def fen_to_yuan(price):

    if price is None:
        return None

    try:
        return round(int(price) / 100, 2)
    except:
        return None





# 定义要保存的Excel文件名
file_name = '价格信息表.xlsx'
#
img_dir = 'output_jpg'
os.makedirs(img_dir, exist_ok=True)
# 准备当前 goods_map 里的新数据，转换为列表字典格式方便后续处理
new_data = []

# 获取 exe 所在目录（打包后）/ 当前脚本目录（开发时）
exe_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))

file_path = os.path.join(exe_dir, file_name)
item_file_path = os.path.join(exe_dir, '商品名称表.xlsx') # strip() 去除可能多输入的空格

if os.path.exists(item_file_path):

    # 读取Excel文件，请将 'your_file.xlsx' 替换为你的实际文件名
    item_df = pd.read_excel(item_file_path)

    # 提取“商品”列的所有名称
    product_names = item_df['商品']

    # 依次处理所有商品名称
    for keyword in product_names:
        pagenum = 1

        # =========================
        # 提取商品价格
        # =========================
        while True:
            params = {
                "param": json.dumps({
                    "terminal": "devtools",
                    "platform": "9",
                    "channel": "miniprograms",
                    "loginId": "a509775c-82fb-46f4-b098-c74e373daa9b",
                    "deviceId": "CB99544E43D000021033F373BAA04D601774569260133",
                    "v": "v6.0.2",
                    "appVersion": "6.0.2",
                    "src": 4,
                    "pageSize": 20,
                    "keyword": keyword,
                    "queryType": 0,
                    "sort": 0,
                    "isOffline": False,
                    "categoryLevel": 0,
                    "fromType": 2,
                    "sortRule": 0,
                    "sortKey": 0,
                    "noResultSearch": 0,
                    "businessCode": 99,
                    "pos": 1,
                    "pageNum": pagenum,
                    "from": 2,
                    "categoryType": 0,
                    "storeInfo": {
                        "venderId": 70121,
                        "defaultChosed": False,
                        "storeId": 721521,
                        "name": "",
                        "businessCode": "99"
                    },
                    "recSceneId": None,
                    "recSceneEdition": "v2",
                    "selectBack": [],
                    "requestSource": "9",
                    "requestVersion": "v6.0.2",
                    "freeVersion": "1.9.2"
                }, ensure_ascii=True),

                "d_track_data": json.dumps({
                    "session_id": "CBA88F4A13000002AFCE52E11D1012EA1778657640792",
                    "project": "metro_mini",
                    "tdc": "",
                    "tpc": "",
                    "uuid": "CB99544E43D000021033F373BAA04D601774569260133",
                    "env": "minip"
                }, ensure_ascii=True)
            }

            response = requests.get(
                url,
                headers=headers,
                params=params,
                timeout=20
            )

            try:
                data = response.json()
            except Exception:
                print(response.text)
                raise

            goods_list = extract_goods(data)


            if not goods_list:
                print("未找到商品数据")
                break

            for item in goods_list:

                promotion = item.get("promotionWareVO") or {}

                tag_price_info = promotion.get("tagPriceInfo") or {}

                img_url = item.get("wareImg") or ""

                sku = item.get("sku")
                name = item.get("wareName", "unknown")

                new_data.append({
                    "name": item.get("wareName"),

                    "sku": item.get("sku"),

                    "current_price": fen_to_yuan(
                        promotion.get("commonPrice")
                    ),

                    "promotion_price": fen_to_yuan(
                        tag_price_info.get("tagPrice")
                    ),

                    "market_price": fen_to_yuan(
                        promotion.get("marketPrice")
                    ),

                    "search_price": fen_to_yuan(
                        item.get("warePrice")
                    ),
                })

                if not img_url:
                    continue


                filename = f"{sku}.jpg"
                filepath = os.path.join(img_dir, filename)

                try:
                    r = requests.get(img_url, timeout=10)
                    r.raise_for_status()

                    with open(filepath, "wb") as f:
                        f.write(r.content)

                except Exception as e:
                    print(f"下载失败: {name}, {e}")

            if pagenum == data["data"]['pageInfo']['pageCount']:
                break
            pagenum += 1


    # =========================
    # 保存文件
    # =========================

    # 将新数据转换为 DataFrame
    new_df = pd.DataFrame(new_data)



    # 判断文件是否存在
    if os.path.exists(file_path):
        print(f"检测到 '{file_path}' 已存在，正在读取并追加数据...")
        # 读取原有的表格数据
        existing_df = pd.read_excel(file_path,engine="openpyxl")
        # 将原有数据与新增数据合并（ignore_index=True 用于重置行索引）
        combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        # 将合并后的完整数据重新写入文件
        combined_df.to_excel(file_path, index=False, engine="openpyxl")
    else:
        print(f"未找到 '{file_path}'，正在新建表格并保存数据...")
        # 直接将新数据写入，创建新文件
        new_df.to_excel(file_path, index=False, engine="openpyxl")



    # =========================
    # 再插入图片
    # =========================

    wb = load_workbook(file_path)
    ws = wb.active

    # 在最前面插入一列
    ws.insert_cols(1)

    # 表头
    ws["A1"] = "图片"

    # 找到 id 列
    id_col = None

    for cell in ws[1]:
        if cell.value == "sku":
            id_col = cell.column
            break

    # 遍历每一行
    for row in range(2, ws.max_row + 1):

        sku = str(ws.cell(row=row, column=id_col).value)

        # 图片路径
        image_path = os.path.join(img_dir, f"{sku}.jpg")

        # 图片存在才插入
        if os.path.exists(image_path):

            img = XLImage(image_path)

            # 设置图片大小
            img.width = 80
            img.height = 80

            # 插入图片
            ws.add_image(img, f"A{row}")

            # 调整行高
            ws.row_dimensions[row].height = 65

    # 调整列宽
    ws.column_dimensions["A"].width = 15

    # 保存
    wb.save(file_path)

    print("Excel生成完成")
else:
    print(f"错误：找不到名为'商品名称表.xlsx'的文件，请检查文件名是否输入正确！")
